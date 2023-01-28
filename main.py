import time
import pandas as pd
import glob
import os
import openpyxl
import  jpype
import json
import  asposecells

################################### FIRST JSON FILE QS SUBJECT 2021 ############################################################

location = "/Users/tripbisagasa/PycharmProjects/QSAutomation/2021 QS Subjects/*.xlsx"
excel_files = glob.glob(location)
writer = pd.ExcelWriter("/Users/tripbisagasa/PycharmProjects/QSAutomation/AllQSSubject2021.xlsx")

for excel_file in excel_files:
    sheet = os.path.basename(excel_file)
    sheet = sheet.split(".")[0]
    df1 = pd.read_excel(excel_file)
    df1['SUBJECT'] = sheet
    df1['YEAR'] = "2021"
    df1.fillna(value="N/A", inplace=True)
    df1.to_excel(writer, sheet_name=sheet, index=False)
writer.save()
print("Done merging with difference sheet inside & Adding Subject & Year")
time.sleep(5)

excel_file = pd.read_excel('AllQSSubject2021.xlsx', sheet_name=None)
dataset_combined = pd.concat(excel_file.values())
dataset_combined.to_excel("SemiFinalSubject2021.xlsx", index=False)
print("Done Combining into 1 sheet")
time.sleep(5)

replace_values = {"51-100":51, "76-100": 76, "101-150":101, "151-200": 151,
                  "201-300":201, "301-400": 301, "201-250":201, "251-300":251,
                  "301-350":301, "351-400":351, "101-120": 101, "401-450":401, "451-500":451,
                  "501-550":501, "551-600":551, "601-650":601, "501-520":501, "201-220": 201,
                  "51-60": 51,"301-320": 301, "51-70": 51, "601-620": 601, "401-410": 401,
                  "201-240": 201, "601-610": 601, "451-460": 451, "301-310": 301, "101-110": 101,
                  "101-140": 101, "151-170": 151, "201-230": 201, "651-670": 651, "101-130": 101,
                  "101-115": 101, "351-370": 351, "551-570": 551, "501-510": 501, "601-630": 601,
                  "301-330": 301, "201-210": 201, "51-80": 51, "351-360": 351, "601-640": 601,
                  "151-160": 151, "301-340": 301, "-":""}

wb = openpyxl.load_workbook("SemiFinalSubject2021.xlsx")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in replace_values.keys():
                cell.value = replace_values.get(cell.value)

wb.save("SemiNewFinal2021.xlsx")
print("Done Replacing values into 1 integer")
time.sleep(3)

jpype.startJVM()
from asposecells.api import Workbook

workbook = Workbook("SemiNewFinal2021.xlsx")
workbook.save("QS Subject 2021.json")
#jpype.shutdownJVM()

print("Done Converting into JSON File")
print("QS Subject 2021 Done")
time.sleep(10)

################################### SECOND JSON FILE QS SUBJECT 2022 ############################################################

location = "/Users/tripbisagasa/PycharmProjects/QSAutomation/2022 QS Subjects/*.xlsx"
excel_files = glob.glob(location)
writer = pd.ExcelWriter("/Users/tripbisagasa/PycharmProjects/QSAutomation/AllQSSubject2022.xlsx")

for excel_file in excel_files:
    sheet = os.path.basename(excel_file)
    sheet = sheet.split(".")[0]
    df1 = pd.read_excel(excel_file)
    df1['SUBJECT'] = sheet
    df1['YEAR'] = "2022"
    df1.fillna(value="N/A", inplace=True)
    df1.to_excel(writer, sheet_name=sheet, index=False)
writer.save()
print("Done merging with difference sheet inside & Adding Subject & Year")
time.sleep(5)

excel_file = pd.read_excel('AllQSSubject2022.xlsx', sheet_name=None)
dataset_combined = pd.concat(excel_file.values())
dataset_combined.to_excel("SemiFinalSubject2022.xlsx", index=False)
print("Done Combining into 1 sheet")
time.sleep(5)

replace_values = {"51-100":51, "76-100": 76, "101-150":101, "151-200": 151,
                  "201-300":201, "301-400": 301, "201-250":201, "251-300":251,
                  "301-350":301, "351-400":351, "101-120": 101, "401-450":401, "451-500":451,
                  "501-550":501, "551-600":551, "601-650":601, "501-520":501, "201-220": 201,
                  "51-60": 51,"301-320": 301, "51-70": 51, "601-620": 601, "401-410": 401,
                  "201-240": 201, "601-610": 601, "451-460": 451, "301-310": 301, "101-110": 101,
                  "101-140": 101, "151-170": 151, "201-230": 201, "651-670": 651, "101-130": 101,
                  "101-115": 101, "351-370": 351, "551-570": 551, "501-510": 501, "601-630": 601,
                  "301-330": 301, "201-210": 201, "51-80": 51, "351-360": 351, "601-640": 601,
                  "151-160": 151, "301-340": 301, "-":""}

wb = openpyxl.load_workbook("SemiFinalSubject2022.xlsx")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in replace_values.keys():
                cell.value = replace_values.get(cell.value)
wb.save("SemiNewFinal2022.xlsx")
print("Done Replacing values into 1 integer")
time.sleep(3)

#jpype.startJVM()
#from asposecells.api import Workbook

workbook = Workbook("SemiNewFinal2022.xlsx")
workbook.save("QS Subject 2022.json")


print("Done Converting into JSON File")
print("QS Subject 2022 Done")
time.sleep(10)

################################### ANALYZE DATA CODE HERE ############################################################

file2021 = open("QS Subject 2021.json", encoding="utf8")
file2022 = open("QS Subject 2022.json", encoding="utf8")

#file2021 = open("SemiNewFinal2021.json", encoding="utf8")
#file2022 = open("SemiNewFinal2022.json", encoding="utf8")

data2021 = json.load(file2021)
data2022 = json.load(file2022)
subjects = []

#loop para kuhaon tanang subjects in distinct
for data in data2021:
    subject = data['SUBJECT']
    if subject not in subjects:
        subjects.append(subject)
for data in data2022:
    subject = data['SUBJECT']
    if subject not in subjects:
        subjects.append(subject)

#sudlanan sa output data
output_data = []

for subject in subjects:
    subject_data_2021 = [data for data in data2021 if data["SUBJECT"] == subject]
    subject_data_2022 = [data for data in data2022 if data["SUBJECT"] == subject]
    for subject2021 in subject_data_2021:
        #mga variables sa 2021 na data
        RANK_2021 = subject2021["RANK"]
        institution = subject2021["UNIVERSITY"]
        year = subject2021["YEAR"]


        #top = data.get("TOP", "UNRANK")
        overall_score_2021 = float(subject2021.get("OVERALL SCORE", 0))
        hindex_citations_2021 = float(subject2021.get("H-INDEX CITATIONS", 0))
        citation_paper_2021 = float(subject2021.get("CITATION PER PAPER", 0))
        academic_reputation_2021 = float(subject2021.get("ACADEMIC REPUTATION", 0))
        employer_reputation_2021 = float(subject2021.get("EMPLOYER REPUTATION", 0))

        #kuhaon data sa subject sa 2022 nga ang institution kay nag match sa current institution sa loop
        data_2022 = [data for data in subject_data_2022 if data["UNIVERSITY"] == institution]

        if len(data_2022) > 0:

            # Kung naa syay 2021 na data
            RANK_2022 = int(data_2022[0]["RANK"])
            new_RANK = RANK_2021 - RANK_2022

            overall_score_2022 = float(data_2022[0].get("OVERALL SCORE", 0))
            new_overall_score = overall_score_2021 - overall_score_2022

            hindex_citation_2022 = float(data_2022[0].get("H-INDEX CITATIONS", 0))
            new_hindex_citation = hindex_citations_2021 - hindex_citation_2022

            citation_paper_2022 = float(data_2022[0].get("CITATION PER PAPER", 0))
            new_citation_pape = citation_paper_2021 - citation_paper_2022

            academic_reputation_2022 = float(data_2022[0].get("ACADEMIC REPUTATION", 0))
            new_academic_reputation = academic_reputation_2021 - academic_reputation_2022

            employer_reputation_2022 = float(data_2022[0].get("EMPLOYER REPUTATION", 0))
            new_employer_reputation = employer_reputation_2021 - employer_reputation_2022


            output_data.append({
                "subject": subject,
                "institution": institution,

                "2021 Rank": RANK_2021,
                "2022 Rank": RANK_2022,
                "Ranking": new_RANK,

                "2021 OVERALL SCORE": overall_score_2021,
                "2022 OVERALL SCORE": overall_score_2022,
                "OVERALL SCORE Difference": new_overall_score,

                "2021 H-INDEX CITATIONS": hindex_citations_2021,
                "2022 H-INDEX CITATIONS": hindex_citation_2022,
                "HINDEX CITATIONS Difference": new_hindex_citation,

                "2021 ACADEMIC REPUTATION": academic_reputation_2021,
                "2022 ACADEMIC REPUTATION": academic_reputation_2022,
                "ACADEMIC REPUTATION Difference": new_academic_reputation,

                "2021 EMPLOYER REPUTATION": employer_reputation_2021,
                "2022 EMPLOYER REPUTATION": employer_reputation_2022,
                "EMPLOYER REPUTATION Difference": new_employer_reputation
            })
        else:
            #Kung wala syay 2022 na data
            output_data.append({
                "subject": subject,
                "institution": institution,
                "2021 RANK": RANK_2021,
                "2022 RANK": RANK_2022,
                "ranking": 0
            })
json_object = json.dumps(output_data, indent=4)

# Writing to sample.json
with open("QS Subject Ranking Test Result.json", "w") as outfile:
    outfile.write(json_object)

print("Done Analyze Converted into JSON")


workbook = Workbook("QS Subject Ranking Test Result.json")
workbook.save("QS Subject Ranking to PBI.xlsx")

jpype.shutdownJVM()

print("Final Result Generated into EXCEL Format (QS Subject Ranking to PBI)")





